import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'https://shobiddak.com/'

response = requests.get(url)
content = response.content

soup = BeautifulSoup(content, 'html.parser')

listings = soup.find_all('div', {'class': 'card white_bg'})

workbook = openpyxl.Workbook()
sheet = workbook.active

headers = ['Price']
for i in range(len(headers)):
    sheet.cell(row=1, column=i+1, value=headers[i])

for i in range(len(listings)):

    price = listings[i].find(
        'div', {'class': 'card-featured-text'}).text

    sheet.cell(row=i+2, column=1, value=price)


workbook.save('products.xlsx')
